#######2022-01-20
import requests
from bs4 import BeautifulSoup as bs
from time import time
import openpyxl ####### 엑셀 write에 사용
import all_companys_list as C ####### 종목 코드 리스트
import test_companys_list as T ####### 구동 테스트를 위한 종목 코드 샘플 리스트
from tqdm import tqdm ####### 진행률 표시 - for문에 사용

####### 필터링 조건
SIGNIFICANT_FIGURE = 3 ## 유효숫자 - ROUND 에서 사용
DIFFERNCE_FILTER = 0 ## 괴리 값(적정가 - 현재가) 필터 - 필터값 이상인 종목 엑셀 저장
EPS_FILTER = 0 ## EPS 필터 - 필터값 이상인 종목 엑셀 저장
PSR_FILTER = 1 ## PSR 필터 - 필터 값 이하인 종목 엑셀 저장
PBR_FILTER = 1 ## PBR 필터 - 필터 값 이하인 종목 엑셀 저장

## MODE 값이 T일 경우 test_companys 참조 C 일 경우(TESTMODE = N or default) companys 참조
## 테스트시 Y / 실사용시 N 
TESTMODE = "N" 
if TESTMODE == "Y":
    MODE = T
elif TESTMODE == "N":
    MODE = C
else:
    MODE = C

####### html 코드 값을 넣어 원하는 위치의 값 추출
def extract(css):
    stockContents = soup.select_one(css)
    TMP = stockContents.text.replace('\n','').strip()
    return TMP.replace(',','').strip()

####### 재무제표에 있는 값을 str로 읽어와 숫자로 변환
def change_float(number):
    try:
        return round(float(number),SIGNIFICANT_FIGURE)
    except ValueError:
        return 0

####### 엑셀 저장을 위한 파트
row = 1
wb = openpyxl.Workbook()
ws = wb.active
####### 컬럼 정의
ws.cell(row=row, column=1).value = "회사명"
ws.cell(row=row, column=2).value = "시가총액"

for company in tqdm(MODE.tickers.keys()):
    ####### AttributeError 발생 시 pass
    try:
        URL = f"https://finance.naver.com/item/main.naver?code={company}"
        req = requests.get(URL)
        html = req.text
        soup = bs(html, "lxml")

        ####### 회사명
        COMPANYNAME = extract("#middle > div.h_company > div.wrap_company > h2 > a")
            
        ####### 시가총액
        CAP = extract("#content > div.section.trade_compare > table > tbody > tr:nth-child(4) > td:nth-child(2)")
        CAP = change_float(CAP)

        ws.cell(row=row+1, column=1).value = COMPANYNAME
        ws.cell(row=row+1, column=2).value = CAP
        row = row + 1
                
    except AttributeError:
        continue
    
print(f"{row} 개의 회사가 저장 되었습니다.")
wb.save("./CAP.xlsx")