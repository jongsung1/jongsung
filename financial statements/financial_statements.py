#######2022-01-20
from asyncio.windows_events import NULL
import requests
from bs4 import BeautifulSoup as bs
from time import time
import datetime
import openpyxl ####### 엑셀 write에 사용
import all_companys_list as C ####### 종목 코드 리스트
import test_companys_list as T ####### 구동 테스트를 위한 종목 코드 샘플 리스트
from tqdm import tqdm ####### 진행률 표시 - for문에 사용
import finance_config as FC ####### config

####### html 코드 값을 넣어 원하는 위치의 값 추출
def extract(css):
    stockContents = soup.select_one(css)
    TMP = stockContents.text.replace('\n','').strip()
    return TMP.replace(',','').strip()

####### 재무제표에 있는 값을 str로 읽어와 숫자로 변환
def change_float(number):
    try:
        return round(float(number),FC.SIGNIFICANT_FIGURE)
    except ValueError:
        return 0

####### extract 함수와 change_float 함수 합
def change(address):
    number = change_float(extract(address))
    try:
        return number
    except ValueError:
        return 0
    except AttributeError:
        return 0

####### 엑셀 저장을 위한 파트
row=1
wb = openpyxl.Workbook()
ws = wb.active

####### 엑셀 저장 함수
def write_excel(COMPANYNAME,CAP,SALES,PROFITS,DEBT,QUICK_RATIO,RESERVE_RATIO,DIVIDEND_RATIO,ROE,PER,PBR,EPS,PSR,PRICE,TARGETPRICE,DIFFERNCE,DIFFERNCE_RATIO,URL,row):
    ws.cell(row=row+1, column=FC.COL_COMPANY).value = COMPANYNAME               ## 회사명
    ws.cell(row=row+1, column=FC.COL_CAP).value = CAP                           ## 시가총액
    ws.cell(row=row+1, column=FC.COL_SALES).value = SALES                       ## 매출
    ws.cell(row=row+1, column=FC.COL_PROFITS).value = PROFITS                   ## 영업이익
    ws.cell(row=row+1, column=FC.COL_DEBT).value = DEBT                         ## 부채
    ws.cell(row=row+1, column=FC.COL_QUICK_RATIO).value = QUICK_RATIO           ## 당좌비율
    ws.cell(row=row+1, column=FC.COL_RESERVE_RATIO).value = RESERVE_RATIO       ## 유보율
    ws.cell(row=row+1, column=FC.COL_DIVIDEND_RATIO).value = DIVIDEND_RATIO     ## 배당률
    ws.cell(row=row+1, column=FC.COL_ROE).value = ROE                           ## ROE
    ws.cell(row=row+1, column=FC.COL_PER).value = PER                           ## PER
    ws.cell(row=row+1, column=FC.COL_PBR).value = PBR                           ## PBR
    ws.cell(row=row+1, column=FC.COL_EPS).value = EPS                           ## EPS
    ws.cell(row=row+1, column=FC.COL_PSR).value = PSR                           ## PSR
    ws.cell(row=row+1, column=FC.COL_PRICE).value = PRICE                       ## 현재가
    ws.cell(row=row+1, column=FC.COL_TARGETPRICE).value = TARGETPRICE           ## 적정가
    ws.cell(row=row+1, column=FC.COL_DIFFERNCE).value = DIFFERNCE               ## 괴리 (적정가 - 현재가)
    ws.cell(row=row+1, column=FC.COL_DIFFERNCE_RATIO).value = DIFFERNCE_RATIO   ## 괴리율% (괴리/현재가)
    ws.cell(row=row+1, column=FC.COL_URL).value = URL                           ## URL

####### 엑셀 헤더
write_excel("회사명","시가총액","매출","영업이익","부채비율(%)","당좌비율(%)","유보율(%)","배당률(%)","ROE","PER","PBR","EPS","PSR","현재가","적정주가","괴리","괴리율(%)","URL",0)

for company in tqdm(FC.MODE.tickers.keys()):
    ####### AttributeError 발생 시 pass
    try:
        URL = f"https://finance.naver.com/item/main.naver?code={company}"
        req = requests.get(URL)
        html = req.text
        soup = bs(html, "lxml")

        ####### 영업이익 - 영업이익 먼저 샘플링, 영업이익이 음수 -> 다음 종목
        PROFITS = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(2) > td.t_line.cell_strong")
        
        ####### 당기 순이익 - 1
        NET_PROFITS1 = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(3) > td:nth-child(2)")
        
        ####### 당기 순이익 - 2
        NET_PROFITS2 = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(3) > td:nth-child(3)")
                
        ####### 당기 순이익 - 3
        NET_PROFITS3 = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(3) > td:nth-child(4)")
        
        ####### 당기 순이익 - 4
        NET_PROFITS4 = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(3) > td:nth-child(5)")

        if FC.FILTER == "Y" or FC.FILTER == "y":
            condition = "PROFITS > 0 and NET_PROFITS1 > 0 and NET_PROFITS2 > 0 and NET_PROFITS3 > 0 and NET_PROFITS4 > 0"
        elif FC.FILTER == "N" or FC.FILTER == "n":
            condition = "PROFITS is not NULL and NET_PROFITS1 is not NULL and NET_PROFITS2 is not NULL and NET_PROFITS3 is not NULL and NET_PROFITS4 is not NULL"

        if condition:

            ####### 회사명
            COMPANYNAME = extract("#middle > div.h_company > div.wrap_company > h2 > a")
            CODE = extract("#middle > div.h_company > div.wrap_company > div > span.code")
            
            ####### 시가총액
            CAP = change("#content > div.section.trade_compare > table > tbody > tr:nth-child(4) > td:nth-child(2)")
            
            ####### 매출
            SALES = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(1) > td.t_line.cell_strong")
                        
            ####### 부채
            DEBT = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(7) > td:nth-child(4)")
            
            ####### 당좌비율
            QUICK_RATIO = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(8) > td:nth-child(4)")
            
            ####### 유보율
            RESERVE_RATIO = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(9) > td:nth-child(4)")

            ####### 배당률
            DIVIDEND_RATIO = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(15) > td:nth-child(4)")

            ####### ROE
            ROE = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(6) > td.t_line.cell_strong")

            ####### EPS
            EPS = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(10) > td.t_line.cell_strong")

            ####### PER
            PER = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(11) > td.t_line.cell_strong")

            ####### PBR
            PBR = change("#content > div.section.cop_analysis > div.sub_section > table > tbody > tr:nth-child(13) > td.t_line.cell_strong")

            ####### PRICE
            PRICE = change("#content > div.section.invest_trend > div.sub_section.right > table > tbody > tr:nth-child(2) > td:nth-child(2) > em")

            ####### PSR
            ####### 매출 값이 없는 회사의 경우 pass
            try:
                PSR = round(CAP/SALES,FC.SIGNIFICANT_FIGURE)
            except ValueError:
                continue
            except ZeroDivisionError:
                continue
        
            ####### 적정 주가
            try:
                TARGETPRICE = ROE*EPS
            except ValueError:
                continue
            
            ####### 괴리 - 적정 주가와 현재 주가의 차이
            DIFFERNCE = TARGETPRICE - PRICE

            ####### 괴리율 - 적정 주가와 현재 주가의 차이
            try:
                DIFFERNCE_RATIO = round(DIFFERNCE/PRICE,FC.SIGNIFICANT_FIGURE)*100
            except ValueError:
                continue
            except ZeroDivisionError:
                continue

            ####### 값들 중 0 인 값이 있는지 확인하기 위해 모두 곱한 값을 변수로 지정
            CHECK_VALUE = SALES*PROFITS*ROE*EPS*PER*PBR*PRICE*PSR*CAP*TARGETPRICE
                        
            if CHECK_VALUE != 0 and FC.FILTER == "Y" or FC.FILTER == "y":
                if DIFFERNCE > FC.DIFFERNCE_FILTER and EPS > FC.EPS_FILTER and QUICK_RATIO > FC.QUICK_RATIO_FILTER:
                    if PSR < FC.PSR_FILTER  and PBR < FC.PBR_FILTER and DEBT < FC.DEBT_FILTER:
                        ####### 엑셀 입력 함수 호출
                        write_excel(COMPANYNAME,CAP,SALES,PROFITS,DEBT,QUICK_RATIO,RESERVE_RATIO,DIVIDEND_RATIO,ROE,PER,PBR,EPS,PSR,PRICE,TARGETPRICE,DIFFERNCE,DIFFERNCE_RATIO,URL,row)
                        row = row + 1
                        
            elif CHECK_VALUE != 0 and FC.FILTER == "N" or FC.FILTER == "n":
                ####### 엑셀 입력 함수 호출
                write_excel(COMPANYNAME,CAP,SALES,PROFITS,DEBT,QUICK_RATIO,RESERVE_RATIO,DIVIDEND_RATIO,ROE,PER,PBR,EPS,PSR,PRICE,TARGETPRICE,DIFFERNCE,DIFFERNCE_RATIO,URL,row)
                row = row + 1
                
    except AttributeError:
        continue
    
print(f"{row-1} 개의 회사가 저장 되었습니다.")
nowtime = datetime.datetime.now()
day = nowtime.strftime("%Y%m%d")
wb.save(f"./financial_statements_{day}.xlsx")